"""
用户清单匹配补全工具
依赖：openpyxl（仅此一个第三方库，打包体积极小）
打包命令：
    pip install pyinstaller openpyxl
    pyinstaller --onefile --windowed --name 用户清单匹配工具 user_list_matcher.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import os

# ── 懒加载 openpyxl，仅在实际读写 xlsx 时导入 ──
def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        messagebox.showerror("缺少依赖", "请先安装 openpyxl：\npip install openpyxl")
        return None


# ─────────────────────── 轻量读写工具函数 ───────────────────────

def read_file(path):
    """读取 xlsx / xls(转csv兼容) / csv，返回 (headers, rows)
    rows: list of dict {列名: 值}
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_csv(path)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return _read_xlsx(path)
    else:
        raise ValueError(f"不支持的文件格式：{ext}")


def _read_csv(path):
    # 自动检测编码
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312", "latin-1"):
        try:
            with open(path, newline="", encoding=enc) as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                rows = [dict(r) for r in reader]
            rows = [{k: (v or "") for k, v in r.items()} for r in rows]
            return list(headers), rows
        except (UnicodeDecodeError, Exception):
            continue
    raise ValueError("无法识别 CSV 文件编码，请另存为 UTF-8 格式后重试。")


def _read_xlsx(path):
    ox = _openpyxl()
    if ox is None:
        return [], []
    wb = ox.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(rows_iter, [])]
    rows = []
    for raw in rows_iter:
        row = {}
        for i, h in enumerate(headers):
            val = raw[i] if i < len(raw) else None
            row[h] = str(val) if val is not None else ""
        rows.append(row)
    wb.close()
    return headers, rows


def write_xlsx(path, headers, rows):
    ox = _openpyxl()
    if ox is None:
        return
    wb = ox.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def write_csv(path, headers, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# ─────────────────────────── 主窗口 ───────────────────────────

class UserListMatcher:
    def __init__(self, root):
        self.root = root
        self.root.title("用户清单匹配补全工具")
        self.root.geometry("920x680")
        self.root.configure(bg="#f0f4f8")
        self.root.resizable(True, True)

        self.full_headers = []
        self.full_rows = []
        self.masked_headers = []
        self.masked_rows = []
        self.mapping_rows = []

        self._build_ui()

    # ──────────────────────── UI 构建 ────────────────────────────

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Microsoft YaHei", 10), padding=6)
        style.configure("TLabel", background="#f0f4f8", font=("Microsoft YaHei", 10))
        style.configure("Header.TLabel", background="#f0f4f8",
                        font=("Microsoft YaHei", 12, "bold"), foreground="#2c3e50")
        style.configure("TLabelframe", background="#f0f4f8")
        style.configure("TLabelframe.Label", background="#f0f4f8",
                        font=("Microsoft YaHei", 10, "bold"), foreground="#34495e")

        main = tk.Frame(self.root, bg="#f0f4f8", padx=16, pady=10)
        main.pack(fill=tk.BOTH, expand=True)

        title_row = tk.Frame(main, bg="#f0f4f8")
        title_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(title_row, text="📋 用户清单匹配补全工具", style="Header.TLabel").pack(side=tk.LEFT)
        tk.Label(title_row, text="Designed by 九龙湖供电所张宽易", bg="#f0f4f8",
                 font=("Microsoft YaHei", 9, "italic"), fg="#95a5a6").pack(side=tk.RIGHT, anchor="e")

        # ── 步骤一：导入文件 ──
        file_frame = ttk.LabelFrame(main, text="第一步：导入文件")
        file_frame.pack(fill=tk.X, pady=(0, 8))
        self._file_row(file_frame, "全量用户清单：", self._load_full, "full")
        self._file_row(file_frame, "脱敏用户清单：", self._load_masked, "masked")

        # ── 步骤二：匹配键 ──
        key_frame = ttk.LabelFrame(main, text="第二步：选择匹配键列（用于关联两份清单的公共标识，如用户编号）")
        key_frame.pack(fill=tk.X, pady=(0, 8))
        ki = tk.Frame(key_frame, bg="#f0f4f8", padx=8, pady=6)
        ki.pack(fill=tk.X)

        ttk.Label(ki, text="全量清单匹配键：").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.full_key_var = tk.StringVar()
        self.full_key_cb = ttk.Combobox(ki, textvariable=self.full_key_var,
                                         state="readonly", width=24)
        self.full_key_cb.grid(row=0, column=1, sticky="w", padx=(0, 24))

        ttk.Label(ki, text="脱敏清单匹配键：").grid(row=0, column=2, sticky="w", padx=(0, 6))
        self.masked_key_var = tk.StringVar()
        self.masked_key_cb = ttk.Combobox(ki, textvariable=self.masked_key_var,
                                           state="readonly", width=24)
        self.masked_key_cb.grid(row=0, column=3, sticky="w")

        # ── 步骤三：列映射 ──
        map_frame = ttk.LabelFrame(main, text="第三步：配置补全列映射")
        map_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        hdr = tk.Frame(map_frame, bg="#dde4ee", padx=8, pady=4)
        hdr.pack(fill=tk.X, padx=8, pady=(6, 0))
        tk.Label(hdr, text="全量清单中取值的列", bg="#dde4ee",
                 font=("Microsoft YaHei", 9, "bold"), width=22, anchor="center").grid(row=0, column=0, padx=(0, 4))
        tk.Label(hdr, text="写入脱敏清单的目标列（留「新增列」则追加）", bg="#dde4ee",
                 font=("Microsoft YaHei", 9, "bold"), width=34, anchor="center").grid(row=0, column=1, padx=(0, 4))
        tk.Label(hdr, text="补全方式", bg="#dde4ee",
                 font=("Microsoft YaHei", 9, "bold"), width=16, anchor="center").grid(row=0, column=2, padx=(0, 4))
        ttk.Button(hdr, text="＋ 添加映射行",
                   command=self._add_mapping_row).grid(row=0, column=3, padx=(8, 0))

        cc = tk.Frame(map_frame, bg="#f0f4f8")
        cc.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        self.map_canvas = tk.Canvas(cc, bg="#f0f4f8", highlightthickness=0)
        sb = ttk.Scrollbar(cc, orient="vertical", command=self.map_canvas.yview)
        self.map_canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.map_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.map_rows_frame = tk.Frame(self.map_canvas, bg="#f0f4f8")
        self._cw = self.map_canvas.create_window((0, 0), window=self.map_rows_frame, anchor="nw")
        self.map_rows_frame.bind("<Configure>",
            lambda e: self.map_canvas.configure(scrollregion=self.map_canvas.bbox("all")))
        self.map_canvas.bind("<Configure>",
            lambda e: self.map_canvas.itemconfig(self._cw, width=e.width))

        # ── 步骤四：输出 & 执行 ──
        bottom = tk.Frame(main, bg="#f0f4f8")
        bottom.pack(fill=tk.X)

        out_frame = ttk.LabelFrame(bottom, text="第四步：输出格式")
        out_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        oi = tk.Frame(out_frame, bg="#f0f4f8", padx=8, pady=6)
        oi.pack(fill=tk.X)
        self.out_fmt_var = tk.StringVar(value="xlsx")
        ttk.Label(oi, text="保存为：").grid(row=0, column=0, sticky="w", padx=(0, 6))
        ttk.Radiobutton(oi, text="Excel (.xlsx)", variable=self.out_fmt_var,
                        value="xlsx").grid(row=0, column=1, sticky="w")
        ttk.Radiobutton(oi, text="CSV (UTF-8)", variable=self.out_fmt_var,
                        value="csv").grid(row=0, column=2, sticky="w", padx=(12, 0))

        run_frame = tk.Frame(bottom, bg="#f0f4f8")
        run_frame.pack(side=tk.RIGHT)
        tk.Button(run_frame, text="▶  开始匹配并导出",
                  font=("Microsoft YaHei", 11, "bold"),
                  bg="#2980b9", fg="white", activebackground="#1a6fa8",
                  relief="flat", padx=18, pady=10,
                  cursor="hand2", command=self._run).pack()

        self.status_var = tk.StringVar(value="请先导入两份清单文件。")
        tk.Label(self.root, textvariable=self.status_var,
                 bg="#2c3e50", fg="#ecf0f1",
                 font=("Microsoft YaHei", 9), anchor="w", padx=10, pady=4
                 ).pack(fill=tk.X, side=tk.BOTTOM)

    def _file_row(self, parent, label, cmd, tag):
        row = tk.Frame(parent, bg="#f0f4f8", padx=8, pady=4)
        row.pack(fill=tk.X)
        ttk.Label(row, text=label, width=14).pack(side=tk.LEFT)
        path_var = tk.StringVar(value="（未选择）")
        setattr(self, f"{tag}_path_var", path_var)
        tk.Label(row, textvariable=path_var, bg="#ffffff", anchor="w",
                 relief="groove", font=("Microsoft YaHei", 9),
                 fg="#555", padx=6, pady=3, width=55).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(row, text="浏览…", command=cmd).pack(side=tk.LEFT)
        info_var = tk.StringVar(value="")
        setattr(self, f"{tag}_info_var", info_var)
        ttk.Label(row, textvariable=info_var, foreground="#27ae60").pack(side=tk.LEFT, padx=8)

    # ──────────────────────── 文件加载 ────────────────────────────

    def _load_file_dialog(self, title):
        return filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel/CSV 文件", "*.xlsx *.xls *.xlsm *.csv"), ("所有文件", "*.*")]
        )

    def _load_full(self):
        path = self._load_file_dialog("选择全量用户清单")
        if not path:
            return
        try:
            headers, rows = read_file(path)
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        self.full_headers, self.full_rows = headers, rows
        self.full_path_var.set(path)
        self.full_info_var.set(f"✓ {len(rows)} 行 × {len(headers)} 列")
        self.full_key_cb["values"] = headers
        if headers:
            self.full_key_cb.current(0)
        self._refresh_mapping_combos()
        self.status_var.set(f"已加载全量清单：{os.path.basename(path)}")

    def _load_masked(self):
        path = self._load_file_dialog("选择脱敏用户清单")
        if not path:
            return
        try:
            headers, rows = read_file(path)
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        self.masked_headers, self.masked_rows = headers, rows
        self.masked_path_var.set(path)
        self.masked_info_var.set(f"✓ {len(rows)} 行 × {len(headers)} 列")
        self.masked_key_cb["values"] = headers
        if headers:
            self.masked_key_cb.current(0)
        self._refresh_mapping_combos()
        self.status_var.set(f"已加载脱敏清单：{os.path.basename(path)}")

    # ──────────────────────── 列映射行 ────────────────────────────

    def _add_mapping_row(self):
        idx = len(self.mapping_rows)
        frame = tk.Frame(self.map_rows_frame, bg="#f0f4f8", pady=2)
        frame.pack(fill=tk.X)

        full_var = tk.StringVar()
        masked_var = tk.StringVar(value="（新增列）")
        mode_var = tk.StringVar(value="overwrite")

        full_cb = ttk.Combobox(frame, textvariable=full_var,
                                values=self.full_headers, state="readonly", width=22)
        full_cb.grid(row=0, column=0, padx=(0, 4))
        if self.full_headers:
            full_cb.current(0)

        masked_vals = ["（新增列）"] + self.masked_headers
        masked_cb = ttk.Combobox(frame, textvariable=masked_var,
                                  values=masked_vals, state="readonly", width=34)
        masked_cb.grid(row=0, column=1, padx=(0, 4))

        mf = tk.Frame(frame, bg="#f0f4f8")
        mf.grid(row=0, column=2, padx=(0, 4))
        tk.Radiobutton(mf, text="覆盖原列", variable=mode_var, value="overwrite",
                       bg="#f0f4f8", font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)
        tk.Radiobutton(mf, text="新增列", variable=mode_var, value="append",
                       bg="#f0f4f8", font=("Microsoft YaHei", 9)).pack(side=tk.LEFT)

        tk.Button(frame, text="✕", fg="#c0392b", bg="#f0f4f8",
                  relief="flat", font=("Microsoft YaHei", 10), cursor="hand2",
                  command=lambda f=frame, i=idx: self._remove_row(f, i)
                  ).grid(row=0, column=3, padx=4)

        self.mapping_rows.append({
            "frame": frame, "full_var": full_var, "masked_var": masked_var,
            "mode_var": mode_var, "full_cb": full_cb, "masked_cb": masked_cb,
            "active": True
        })

    def _remove_row(self, frame, idx):
        frame.destroy()
        if idx < len(self.mapping_rows):
            self.mapping_rows[idx]["active"] = False

    def _refresh_mapping_combos(self):
        masked_vals = ["（新增列）"] + self.masked_headers
        for r in self.mapping_rows:
            if not r["active"]:
                continue
            r["full_cb"]["values"] = self.full_headers
            r["masked_cb"]["values"] = masked_vals

    # ──────────────────────── 执行匹配 ────────────────────────────

    def _run(self):
        if not self.full_rows or not self.masked_rows:
            messagebox.showwarning("提示", "请先导入全量清单和脱敏清单！")
            return

        full_key = self.full_key_var.get().strip()
        masked_key = self.masked_key_var.get().strip()
        if not full_key or not masked_key:
            messagebox.showwarning("提示", "请选择匹配键列！")
            return

        active = [r for r in self.mapping_rows if r["active"]]
        if not active:
            messagebox.showwarning("提示", "请至少添加一条列映射关系！")
            return

        # 构建查找表
        lookup = {}
        for row in self.full_rows:
            k = str(row.get(full_key, "")).strip()
            if k:
                lookup[k] = row

        # 确定输出列顺序，处理新增列命名冲突
        result_headers = list(self.masked_headers)
        col_plan = []  # [(src_col, write_col)]
        for r in active:
            src_col = r["full_var"].get()
            dst_col = r["masked_var"].get()
            mode = r["mode_var"].get()
            if not src_col:
                continue
            if dst_col == "（新增列）" or mode == "append":
                base = src_col
                write_col = base
                suffix = 1
                while write_col in result_headers:
                    write_col = f"{base}_补全{suffix}"
                    suffix += 1
                result_headers.append(write_col)
            else:
                write_col = dst_col
            col_plan.append((src_col, write_col))

        # 生成结果行
        result_rows = []
        matched_count = 0
        for row in self.masked_rows:
            new_row = dict(row)
            key_val = str(row.get(masked_key, "")).strip()
            src = lookup.get(key_val)
            if src:
                matched_count += 1
            for src_col, write_col in col_plan:
                new_row[write_col] = src.get(src_col, "") if src else ""
            result_rows.append(new_row)

        total = len(result_rows)
        unmatched = total - matched_count

        out_fmt = self.out_fmt_var.get()
        ext = ".xlsx" if out_fmt == "xlsx" else ".csv"
        out_path = filedialog.asksaveasfilename(
            title="保存结果文件",
            defaultextension=ext,
            filetypes=[("Excel 文件", "*.xlsx"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not out_path:
            return

        try:
            if out_fmt == "xlsx":
                write_xlsx(out_path, result_headers, result_rows)
            else:
                write_csv(out_path, result_headers, result_rows)

            msg = (f"✅ 匹配完成！\n\n"
                   f"总行数：{total}\n"
                   f"成功匹配：{matched_count} 行\n"
                   f"未匹配（键值不存在）：{unmatched} 行\n\n"
                   f"已保存至：\n{out_path}")
            messagebox.showinfo("完成", msg)
            self.status_var.set(
                f"已导出：{os.path.basename(out_path)}  | 匹配 {matched_count}/{total} 行")
        except Exception as e:
            messagebox.showerror("保存失败", f"保存文件时出错：\n{e}")


# ─────────────────────────── 入口 ───────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = UserListMatcher(root)
    root.mainloop()
